QUIT = 0

try:
    import openpyxl
except ImportError:
    print("Please install the openpyxl module (pip install openpyxl)")
    QUIT = 1

try:
    from rapidfuzz import fuzz
except ImportError:
    print("Please install the rapidfuzz module (pip install rapidfuzz)")
    QUIT = 1

import multiprocessing as mp
# mp.set_start_method('spawn')
import csv
import re
orig_filename = ""


class item:
    # Internal vars to represent an item.
    # Name: Name of an item
    # ID: InfoGenesis Item ID
    # Price: InfoGenesis Price
    # Inserted: Internal var that gets set when placed in the spreadsheet

    def __init__(self, name, rep, category):
        self.name = name
        self.inserted = 0
        self.rep = rep
        self.category = category


def import_items():
    global orig_filename
    orig_filename = input("Please enter file path to MI_EXP file: ")
    items = []
    counter = 0
    with open(orig_filename) as f:
        lines = f.readlines()
        for i in range(0, len(lines)):
            # Split, basic parsing, store into object
            # Standardize line because of brackets
            # This part handles the sku info
            try:
                # So, this isn't guaranteed to have a match, in which case we're going to get an IndexError
                # for a list that doesn't exist. This basically says "if we have a match, cool. If not, oh well"
                skuList = re.findall(r"(\{\"(?:.*)\"\})", lines[i])[0]
                passOne = re.sub(r"(\{\"(?:.*)\"\})", skuList.replace('"', ''), lines[i])
            except:
                passOne = lines[i]
            entry = ""
            for char in passOne:
                if char == '{':
                    entry += '"'
                    entry += char
                elif char == '}':
                    entry += char
                    entry += '"'
                else:
                    entry += char
            entry = [ "{}".format(x) for x in next(csv.reader([entry], delimiter=',', quotechar='"')) ]
            object = item(entry[2].strip("\""), entry, int(entry[8]))
            items.append(object)
            counter += 1
            print(f"Parsed {counter} items from export file...", end='\r')
            # print(f"added {object.name} for revenue class {category}!")

    # We should have a list of all items with some info
    print()
    return items


def split_items(items):
    igCategories = []
    counter = 0
    for category in range(0, 50):
        catItems = []
        for item in items:
            if item.category == category:
                catItems.append(item)
        if len(catItems) > 0:
            # We have items to actually add
            igCategories.append(catItems)
            counter += 1
        print(f"Split into {counter} categories...", end='\r')

    print()
    return igCategories


def outList(items):
    '''
    Short helper function that will output
    all of the items in a list to a output text file
    to aid in debugging
    '''

    with open("output.txt", "w") as f:
        for item in items:
            f.write(",".join(item.rep) + "\n")

    print("outList complete!")
    quit()


def sort_items(todo, counter, results):
    # Agh. Suuuuper inefficient
    sorted_list = []
    unsorted = todo.get()
    allItems = unsorted.copy()
    for item in unsorted:
        if item.inserted != 1:
            # Need to add this and similar items to the sorted list
            sorted_list.append(item)
            item.inserted = 1
            counter.put(1)
            for item1 in allItems:
                # This is about o(N^2). Yikes.
                if item1.inserted != 1:
                    # Not in the list, is it similar?
                    if fuzz.token_sort_ratio(item.name, item1.name) >= 80:
                        sorted_list.append(item1)
                        item1.inserted = 1
                        counter.put(1)

    # we should be done with this block now
    results.put(sorted_list)
    counter.put(None)
    return


def getCategory(item):
    return item.category


def init_sort(igCategories):
    # This actually spawns the sort threads (processes in this case)
    sorted_list = []
    processPool = mp.Pool()
    poolManager = mp.Manager()
    todo = poolManager.Queue()
    sorted_results = poolManager.Queue()
    counter_queue = poolManager.Queue()
    counter = 0
    workers = len(igCategories)

    # Add tasks to todo (all lists to be sorted)
    for items in igCategories:
        todo.put(items)

    items = poolManager.list()
    for i in igCategories:  # Make proper var names
        for j in i:
            items.append(j)

    for i in igCategories:
        res = processPool.apply_async(sort_items, (todo, counter_queue, sorted_results))

    while workers > 0:  # Change this for number of results expected
        status = counter_queue.get()
        if status == None:
            workers -= 1
        else:
            counter += status  # this increments the counter for every item sorted
            progress = "{:.2f}".format((counter / len(items)) * 100)
            print(f"Sorting Progress: {progress}%", end='\r')
    print()

    processPool.close()
    processPool.join()

    # We should have results
    while not sorted_results.empty():
        result = sorted_results.get()
        for item in result:
            sorted_list.append(item)

    # sorted_list.sort(key=getCategory)
        
    

    return sorted_list


##############################################################################################
# SORTING CODE ABOVE
# SPREADSHEET CODE BELOW
##############################################################################################

def create_spreadsheet(items):
    output_file = openpyxl.Workbook()
    output = output_file.create_sheet("Item Export")
    sku_out = output_file.create_sheet("SKU List")
    output_file.remove(output_file.active)  # Deletes default sheet
    main_counter = 3
    sku_counter = 4

    # Initial Formatting Setup
    output.cell(row = 2, column = 2).value = "UD Item"
    output.cell(row = 2, column = 3).value = "UD Name"
    output.cell(row = 2, column = 4).value = "Price"
    output.cell(row = 2, column = 5).value = "Revenue Category"
    output.cell(row = 2, column = 6).value = "SKU"
    sku_out.cell(row = 3, column = 2).value = "UD Item"
    sku_out.cell(row = 3, column = 3).value = "SKUs Associated"

    for item in items:
        # Let's first populate the easy things, starting with the item number and name
        output.cell(row = main_counter, column = 2).value = str(item.rep[1])
        output.cell(row = main_counter, column = 3).value = item.rep[2]

        # Now to get the price, we will need to do some parsing
        price_array = item.rep[6].split(",")
        price = price_array[1].strip("}")

        output.cell(row = main_counter, column = 4).value = price

        # Revenue Category
        output.cell(row = main_counter, column = 5).value = str(item.category)

        # Now... SKUs. We need a few rules for these and some parsing
        sku_array = item.rep[14].strip("{}").split(",")
        # print(item.rep)
        if len(sku_array) <= 2:
            # We have one SKU, goes in original sheet
            sku = sku_array[0]
            output.cell(row = main_counter, column = 6).value = sku
        else:
            # We have more than one SKU, need to use other sheet
            output.cell(row = main_counter, column = 6).value = "Multiple"
            # Let's make a hyperlink between the two cells so we can reference them easily
            output.cell(row = main_counter, column = 6).hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(ref="F" + str(main_counter), location="'SKU List'!B" + str(sku_counter), tooltip=None, display='Multiple', id=None)
            output.cell(row = main_counter, column = 6).font = openpyxl.styles.Font(color="000000FF", underline='single')
            sku_out.cell(row = sku_counter, column = 2).value = str(item.rep[1])
            sku_out.cell(row = sku_counter, column = 2).hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(ref="B" + str(sku_counter), location="'Item Export'!F" + str(main_counter), tooltip=None, display=str(item.rep[1]), id=None)
            sku_out.cell(row = sku_counter, column = 2).font = openpyxl.styles.Font(color="000000FF", underline='single')
            for sku in range(0, len(sku_array), 2):
                sku_out.cell(row = sku_counter, column = 3).value = sku_array[sku]
                sku_counter += 1
            sku_counter += 1  # Adds a blank row after an item, adds some separation between items
            
        main_counter += 1
        progress = "{:.0f}".format((main_counter / len(items)) * 100)
        print(f"{progress}% of sheet created.", end='\r')

    # The sheet should be filled at this point, let's format it as a table now!
    defaultStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleLight1', showFirstColumn=False, showLastColumn=False, showColumnStripes=False, showRowStripes=True)
    tab1 = openpyxl.worksheet.table.Table(displayName="Items", ref="B2:F" + str(output.max_row))
    tab2 = openpyxl.worksheet.table.Table(displayName="skus", ref="B3:C" + str(sku_out.max_row))
    tab1.tableStyleInfo = defaultStyle
    tab2.tableStyleInfo = defaultStyle
    output.add_table(tab1)
    sku_out.add_table(tab2)
    
    # Let's resize the columns so we don't need to do that in excel
    for column_cells in output.columns:
        length = max(len(cell.value or "") for cell in column_cells) + 2.65
        output.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length
    for column_cells in sku_out.columns:
        length = max(len(cell.value or "") for cell in column_cells) + 2.65
        sku_out.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length
        
    output_file.save("MI_Exp_Converted_" + orig_filename.split("_")[2].strip(".txt") + ".xlsx")
    print()
    

def main():
    if QUIT == 1:
        quit()
    items = import_items()
    # outList(items)
    igCategories = split_items(items)
    sorted = init_sort(igCategories)
    create_spreadsheet(sorted)


if __name__ == "__main__":
    main()
